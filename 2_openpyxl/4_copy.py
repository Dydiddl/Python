import openpyxl
from datetime import datetime

# A 파일 불러오기
input_wb = openpyxl.load_workbook('일용직 지명원_NewVersion.xlsx')
input_ws = input_wb.active

# B 파일 생성하기
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# A 의 내용을 B로 복사하기
for row in input_ws.iter_rows(values_only=True):
    output_ws.append(row)

# 번호 매김(1행 해더)
for row in range(2, output_ws.max_row) :
    if output_ws[f'B{row}'].value:
        output_ws[f'A{row}'] = row - 1
        
# 주민등록번호로 나이 계산하기
for row in range (2, output_ws.max_row):
    jumin = output_ws[f'C{row}'].value
    if jumin:
        birth_year = int(jumin[:2])
        birth_code = int(jumin[7]) # 주민등록번호 뒷자리 첫 번째 숫자
        
        # 주민등록번호 뒷자리 첫 번째 숫자에 따라 20세기 출생자 또는 21세기 출생자로 구분
        if birth_code in [1, 2]:
            birth_year += 1900
        elif birth_code in [3, 4]:
            birth_year += 2000
            
        birth_month = int(jumin[2:4])
        birth_day = int(jumin[4:6])
        
        birth_date = datetime(birth_year, birth_month, birth_day)
        
        today = datetime.today()
        age = today.year - birth_year - ((today.month, today.day) < (birth_month, birth_day))
        
        output_ws[f'D{row}'] = age        
        
# 주민등록번호가 없을 경우 '주민등록번호 없음' 입력
for row in range(2, output_ws.max_row + 1):
    if not output_ws[f'C{row}'].value:
        output_ws[f'D{row}'] = '민증없음'
        

# 변경사항을 저장
output_wb.save('일용직 지명원_add.xlsx')
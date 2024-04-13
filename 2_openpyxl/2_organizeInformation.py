import openpyxl

wb = openpyxl.load_workbook('정리필요 일용직 지명원_나이계산.xlsx')

input_ws = wb['Sheet']
output_ws = wb.create_sheet('Sheet2')

rows_list = []

for row in input_ws.iter_rows():
    row_values = []
    for cell in row:
        row_values.append(cell.value)
    rows_list.append(row_values)

with_jumin = []
without_jumin = []

for item in rows_list:
    if item[2] is not None:
        with_jumin.append(item)
    else:
        without_jumin.append(item)

for row_idx, row_data in  enumerate(with_jumin, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        output_ws.cell(row=row_idx, column=col_idx, value=value)

for row_idx, row_data in  enumerate(without_jumin, start=len(with_jumin) + 1):
    for col_idx, value in enumerate(row_data, start=1):
        output_ws.cell(row=row_idx, column=col_idx, value=value)


# A 열에 순서대로 번호 매김
for row in range(2, output_ws.max_row + 1) :
    if output_ws[f'B{row}'].value:
        output_ws[f'A{row}'] = row - 1


# sheet1을 지우기
# wb.remove_sheet(sheet)
# print(with_jumin)
# print(without_jumin)
    
wb.save('정리필요 일용직 지명원_나이계산_민증필요.xlsx')
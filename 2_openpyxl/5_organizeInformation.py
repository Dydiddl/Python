import openpyxl

wb = openpyxl.load_workbook('일용직 지명원_add.xlsx')

ws_a = wb['Sheet']
ws_b = wb.create_sheet('Sheet2')

rows_list = []

for row in ws_a.iter_rows():
    row_values = []
    for cell in row:
        row_values.append(cell.value)
    rows_list.append(row_values)

with_jumin = []
without_jumin = []

for item in rows_list:
    if item[2] is not None:
        with_jumin.append(item)
    else:
        without_jumin.append(item)

for row_idx, row_data in  enumerate(with_jumin, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        ws_b.cell(row=row_idx, column=col_idx, value=value)

for row_idx, row_data in  enumerate(without_jumin, start=len(with_jumin) + 1):
    for col_idx, value in enumerate(row_data, start=1):
        ws_b.cell(row=row_idx, column=col_idx, value=value)


# A 열에 순서대로 번호 매김
for row in range(2, ws_b.max_row + 1) :
    if ws_b[f'B{row}'].value:
        ws_b[f'A{row}'] = row - 1


# sheet1을 지우기
# wb.remove_sheet(sheet)
# print(with_jumin)
# print(without_jumin)
    
wb.save('일용직 지명원_iter_rows_test.xlsx')
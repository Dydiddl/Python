import openpyxl

wb = openpyxl.load_workbook('일용직 지명원.xlsx')
ws = wb.active

for row in range(2, ws.max_row + 1) :
    if ws[f'B{row}'].value:
        ws[f'A{row}'] = row - 1

wb.save('일용직 지명원_newVersion.xlsx')